# -*- coding: utf-8 -*-
"""DEPLOY_DATA_DRUGS_TYPE.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1do9v2ntDYlZ3_aO5PfTXNNPf85aYz1KL

#DEPLOYMENT
"""
import streamlit as st
import streamlit as st,requests
import pandas as pd
import numpy as np
import time
import matplotlib.pyplot as plt
import plotly.figure_factory as ff
from sklearn.model_selection import train_test_split
import plotly.express as px
import seaborn as sns
from sklearn.model_selection import GridSearchCV
import types
from openpyxl import load_workbook
import openpyxl
import os
import seaborn as sns

from sklearn.preprocessing import OrdinalEncoder
from sklearn.metrics import accuracy_score,recall_score,precision_score,f1_score,confusion_matrix,roc_curve,roc_auc_score
from sklearn.preprocessing import StandardScaler
from sklearn.tree import DecisionTreeClassifier

from sklearn.exceptions import NotFittedError
from scipy.stats import randint
import missingno as msng

from sklearn.model_selection import cross_val_score,KFold

from collections import Counter
from imblearn.under_sampling import RandomUnderSampler
from imblearn.over_sampling import SMOTE

st.header("DRUG_DATA_SET")
st.image("https://media.istockphoto.com/id/692096736/photo/concrete-pouring-during-commercial-concreting-floors-of-building.jpg?s=1024x1024&w=is&k=20&c=XYYH7UhgqsMmwGBWO6UJsxaSgjxNDuQO8i7N27nwRlk=", width=200)
def DATA_FRAME(df):
  url_1= 'https://raw.githubusercontent.com/merrooo/ML_DEPLOY_DRUGS/main/DRUGS.csv?token=GHSAT0AAAAAACNFDRMB6H7UC5FLSIO6RE7YZNNUL2A'
  df=pd.read_csv(url_1)
  return df
page=st.sidebar.selectbox("PROSSECCES_ON_DATA_SET",("- -","- EDA -","- VISUALIZATION -","- PREDICTION -"))
if page == "- EDA -":
    EDA_=st.sidebar.selectbox("EXPLORING_DATA_COLUMNS",("- -","- DATA_FARME -","- COLUMNS -"))

    if EDA_ == "- DATA_FARME -":
      EXPLORE_DATA_=st.sidebar.toggle('EXPLORE_DATA',disabled=False)
      CHECK_NULL_=st.sidebar.toggle('CHECK_NULL',disabled=False)
      DESCRIBE_=st.sidebar.toggle('DESCRIBE',disabled=False)
      COLUMNS_=st.sidebar.toggle('COLUMNS',disabled=False)
      MAX_=st.sidebar.toggle('MAXIMUM_VALUES_FEATURES',disabled=False)
      HEAD_=st.sidebar.toggle('DATA_HEAD',disabled=False)
      if EXPLORE_DATA_:
        st.dataframe(DATA_FRAME('df'))
        st.write('DONE!!')
      elif MAX_:
        st.write('MAXIMUM_VALUES_FEATURES!!')
        data_matrix = [['ITEM', 'DESCIBTION'],
               ['Age',74],
               ['Na_to_K', 38.247]]
        fig = ff.create_table(data_matrix)
        st.plotly_chart(fig)
        st.write('DONE!!')
      elif HEAD_:
        st.write('DATA_HEAD!!')
        st.dataframe(DATA_FRAME('df').head())
        st.write('DONE!!')
      elif CHECK_NULL_:
        st.write('EXPLORING_NULL!!')
        st.dataframe(DATA_FRAME('df').isnull().sum())
        st.write('DONE!!')
      elif DESCRIBE_:
        st.write('DESCRIPE_OF_DATA!!')
        st.dataframe(DATA_FRAME('df').describe())
        st.write('DONE!!')
      elif COLUMNS_:
        st.write('EXPLORE_HOW_MANY_FEATURES!!')
        st.dataframe(DATA_FRAME('df').columns)
        st.write('DONE!!')
    elif EDA_ == "- COLUMNS -":
      st.write('VALUE_COUNTS_FOR_OUT_PUT[Drug]!!')
      VALUE_COUNTS_=st.sidebar.toggle('VALUE_COUNTS',disabled=False)
      UNIQUE_=st.sidebar.toggle('UNIQUE',disabled=False)
      if VALUE_COUNTS_:
        st.write('CHECK_FOR_OUTPUT_[Drug]!!')
        Drug_=st.toggle('Drug',disabled=False)
        if Drug_:
         st.dataframe(DATA_FRAME('df')['Drug'].value_counts())
         st.write('DONE!!')
      elif UNIQUE_:
        st.write('CHECK_FOR_OUTPUT_[Drug]!!')
        Drug_=st.toggle('Drug',disabled=False)
        if Drug_:
         st.dataframe(DATA_FRAME('df')['Drug'].unique())
         st.write('DONE!!')
    st.balloons()
#_______________________________________________________________________________________________________________________________________________________________

elif page =="- VISUALIZATION -":
   #------------------------------------------------------------------
    st.sidebar.write('AREA_CHART!!')
    button_VISU_2=st.sidebar.button("AREA_CHART",type="primary")
    st.sidebar.write('DISTRIBUTION_PLOTTING!!')
    if button_VISU_2:
     st.write("AREA_CHART_DEPENDENT FEATURES (CEMENT_AGE_SUPERPLASTICIZER) AFFECTING ON STRENGTH")
     st.area_chart(
     DATA_FRAME('df'), x="Strength", y=["Cement", "Superplasticizer","Age"], color=["#f0e936", "#4633f2","#0e6210"]) # Optional
     st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_3=st.sidebar.button("DISTRUBTION_PLOT_1",type="primary")
    if button_VISU_3:
      st.write("DISTRIBUTION_PLOTTING THE DEPENDENT FEATURES (CEMENT_AGE_SUPERPLASTICIZER) AFFECTING ON STRENGTH")
      x1 = np.random.randn(200) - 2
      x2 = np.random.randn(200)
      x3 = np.random.randn(200) + 2
      hist_data = [x1, x2, x3]
      group_labels = ['Cement', 'Age', 'Superplasticizer']
      fig = ff.create_distplot(
        hist_data, group_labels, bin_size=[.1, .25, .5])
      st.plotly_chart(fig, use_container_width=True)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_4=st.sidebar.button("DISTRUBTION_PLOT_2",type="primary")
    if button_VISU_4:
      st.write("DISTRIBUTION_PLOTTING CEMENT REGARDING TO STRENGTH ")
      x1 = np.random.randn(1000) - 2
      x2 = np.random.randn(1000)
      x3 = np.random.randn(1000) + 2
      hist_data = [x3]
      group_labels = ['Strength']
      fig = ff.create_distplot(
          hist_data, group_labels, bin_size=[.1, .25, .5])
      st.plotly_chart(fig, use_container_width=True)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_5=st.sidebar.button("DISTRUBTION_PLOT_3",type="primary")
    if button_VISU_5:
     st.write("DISTRIBUTION_PLOTTING CEMENT REGARDING TO STRENGTH ")
     x1 = np.random.randn(1000) - 2
     x2 = np.random.randn(1000)
     x3 = np.random.randn(1000) + 2
     hist_data = [x3]
     group_labels = ['Cement']
     fig = ff.create_distplot(
          hist_data, group_labels, bin_size=[.1, .25, .5])
     st.plotly_chart(fig, use_container_width=True)
     st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_6=st.sidebar.button("DISTRUBTION_PLOT_4",type="primary")
    if button_VISU_6:
      st.write("DISTRIBUTION_PLOTTING CEMENT REGARDING TO STRENGTH ")
      x1 = np.random.randn(1000) - 2
      x2 = np.random.randn(1000)
      x3 = np.random.randn(1000) + 2
      hist_data = [x3]
      group_labels = ['Age']
      fig = ff.create_distplot(
         hist_data, group_labels, bin_size=[.1, .25, .5])
      st.plotly_chart(fig, use_container_width=True)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_7=st.sidebar.button("DISTRUBTION_PLOT_5",type="primary")
    st.sidebar.write('HEAT_MAP!!')
    if button_VISU_7:
      st.write("DISTRIBUTION_PLOTTING CEMENT REGARDING TO STRENGTH ")
      x1 = np.random.randn(1000) - 2
      x2 = np.random.randn(1000)
      x3 = np.random.randn(1000) + 2
      hist_data = [x3]
      group_labels = ['Superplasticizer']
      fig = ff.create_distplot(
          hist_data, group_labels, bin_size=[.1, .25, .5])
      st.plotly_chart(fig, use_container_width=True)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_8=st.sidebar.button("HEAT_MAP_DF",type="primary")
    if button_VISU_8:
      st.write("HEAT_MAP(DATA_FRAME)")
      fig = px.imshow(DATA_FRAME('df'), text_auto=True, aspect="auto")
      st.plotly_chart(fig)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_9=st.sidebar.button("HEAT_MAP_CEMENT",type="primary")
    if button_VISU_9:
      st.write("HEAT_MAP(CEMENT_STRENGTH)")
      fig = px.density_heatmap(DATA_FRAME('df'), x="Cement", y="Strength", text_auto=True, nbinsx=7, color_continuous_scale='turbid_r', width=686, height=889)
      st.plotly_chart(fig)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_10=st.sidebar.button("HEAT_MAP_AGE",type="primary")
    if button_VISU_10:
      st.write("HEAT_MAP(AGE_STRENGTH)")
      fig = px.density_heatmap(DATA_FRAME('df'), x="Age", y="Strength", text_auto=True, nbinsx=7, color_continuous_scale='turbid_r', width=686, height=889)
      st.plotly_chart(fig)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_11=st.sidebar.button("HEAT_MAP_Superplasticizer",type="primary")
    if button_VISU_11:
      st.write("HEAT_MAP(Superplasticizer_STRENGTH)")
      fig = px.density_heatmap(DATA_FRAME('df'), x="Superplasticizer", y="Strength", text_auto=True, nbinsx=7, color_continuous_scale='turbid_r', width=686, height=889)
      st.plotly_chart(fig)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_12=st.sidebar.button("HEAT_MAP_Fly Ash_STRENGTH",type="primary")
    if button_VISU_12:
      st.write("HEAT_MAP(Fly Ash_STRENGTH)")
      fig = px.density_heatmap(DATA_FRAME('df'), x="Fly Ash", y="Strength", text_auto=True, nbinsx=7, color_continuous_scale='turbid_r', width=686, height=889)
      st.plotly_chart(fig)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")

    button_VISU_13=st.sidebar.button("HEAT_MAP_Water",type="primary")
    if button_VISU_13:
      st.write("HEAT_MAP(Superplasticizer_STRENGTH)")
      fig = px.density_heatmap(DATA_FRAME('df'), x="Water", y="Strength", text_auto=True, nbinsx=7, color_continuous_scale='turbid_r', width=686, height=889)
      st.plotly_chart(fig)
      st.success('ALREADY_GRAPH_VISUALIZED!', icon="✅")
    st.balloons()
#_______________________________________________________________________________________________________________________________________________________________

elif page =="- PREDICTION -":
  st.write("WE_NEED_SOME_INFORMATION_TO_PREDICT_THE DRUG_TYPE")
  #------------------------------------------------------------------
  st.write('DATA_HEAD!!')
  url_1= 'https://raw.githubusercontent.com/merrooo/ML_DEPLOY_DRUGS/main/DRUGS.csv?token=GHSAT0AAAAAACNFDRMB6H7UC5FLSIO6RE7YZNNUL2A'
  df=pd.read_csv(url_1)
  oe = OrdinalEncoder(categories=[['DrugY', 'drugC', 'drugX', 'drugA', 'drugB']])
  df['Drug'] = oe.fit_transform(df[['Drug']])
  df= pd.get_dummies(DATA_FRAME('df'), columns=['BP', 'Cholesterol','Sex'])
  st.dataframe(df.head(5))
  ok=st.button("PREDICTION_DRUGS_TYPE")

  if ok:

    oe = OrdinalEncoder(categories=[['DrugY', 'drugC', 'drugX', 'drugA', 'drugB']])
    df['Drug'] = oe.fit_transform(df[['Drug']])

    df = pd.get_dummies(df, columns=['BP', 'Cholesterol','Sex'])

    x = DATA_FRAME('df').loc[:, DATA_FRAME('df').columns != 'Drug']
    y = DATA_FRAME('df')['Drug']

    Rus = RandomUnderSampler(sampling_strategy = {0:16, 1:16, 2:16 , 3:16, 4:16},random_state=42)
    x_RUS, y_RUS= Rus.fit_resample(x, y)

    scaler=StandardScaler()
    x_train=scaler.fit_transform(x_train)
    x_test=scaler.transform(x_test)

    x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=.3, random_state=42)

    BAGGING_CLAS_mode=BaggingClassifier()
    kf=KFold(n_splits=5,shuffle=True,random_state=0)
    score=cross_val_score(BAGGING_CLAS_mode,x_train,y_train,cv=kf) # kfold
    params= {'n_estimators' : [10, 100, 1000]}
    grid_search_BAGG=GridSearchCV(
    estimator=BAGGING_CLAS_mode,
    param_grid=params,verbose = 1, n_jobs = -1,
    scoring='accuracy',cv=kf)
    grid_result=grid_search_BAGG.fit(x_train,y_train)

    DRUG_ = BAGGING_CLAS_mode.predict(x_test) # Uses the model to predict the drug type
    st.subheader(f" THE_ESTIMATED_DRUG_TYPE_IS :- \n[{DRUG_[0]:.2f}]") # Displays the predicted drug type in Streamlit

    new_data=pd.DataFrame(n,columns=['Age_','Na_to_K_','CHOLESTROL_'])
    new_data['DRUG_'] = DRUG_

    st.subheader('PREDICTION_SAMPLE')
    st.dataframe(new_data)

    progress_text = "Operation in progress. Please wait."
    my_bar = st.progress(0, text=progress_text)

    for percent_complete in range(100):
        time.sleep(0.01)
        my_bar.progress(percent_complete + 1, text=progress_text)
    time.sleep(1)
    my_bar.empty()

    #--------------------------------------------------------------------------
    new_data['DRUG_'] = DRUG_
    st.write('------------------------------ACCURACY_TRAIN-----------------------------')
    DRUG_TRAIN=XGB_REG_model.predict(x_train)
    SCORE_TRAIN=r2_score(y_train,DRUG_TRAIN)*100
    st.subheader(" ACCURACY_TRAIN_FOR_MODEL_IS :- \n[{:.2f}] %".format(SCORE_TRAIN))
    st.write('------------------------------ACCURACY_TEST------------------------------')
    DRUG_TEST=XGB_REG_model.predict(x_test)
    SCORE_TEST=r2_score(y_test,DRUG_TEST)*100
    st.subheader(" ACCURACY_TEST_FOR_MODEL_IS :- \n[{:.2f}] %".format(SCORE_TEST))
    st.write('-----------------------------ACCURACCY_GRAPH----------------------------')
    labels = 'ACCURACY_TEST', 'ACCURACY_TRAIN'
    sizes = [48.6, 51.4]
    explode = (0, 0.1)
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
          shadow=True, startangle=90)
    ax1.axis('equal')
    st.pyplot(fig1)
    st.write('WOREST_RESULT_FOR_DRUG_TYPE')
    data_matrix = [['ITEM', 'DESCIBTION'],
               ['Na_to_K_',38.247]]
    fig = ff.create_table(data_matrix)
    st.plotly_chart(fig)
    #--------------------------------------------------------------------------
    st.title('EXPORT_PREDICTION_SAMPLE_AS_CSV.FILE')
    def my_hash_func(value):
        return id(value)
    @st.cache(hash_funcs={types.FunctionType: my_hash_func})
    def convert_df(new_data):
        return new_data.to_csv().encode('utf-8')
# Create a new workbook if 'large_df.xlsx' doesn't exist
    # Create a new workbook if 'large_df.xlsx' doesn't exist
    if not os.path.exists('large_df.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('large_df.xlsx')
    wb = openpyxl.load_workbook('large_df.xlsx')
    ws = wb.active
    for index, row in new_data.iterrows():
        ws.append(row.tolist())
    wb.save('large_df.xlsx')
    csv = convert_df(new_data)
    st.download_button(
        label="DOWNLOAD_CSV.FILE",
        data=csv,
        file_name='large_df.csv',
        mime='text/csv')
    st.write('PREDICTION_DATA_ADD_SUCCESSFULLY_TO_CSV.FILE')
else:

  if 'sidebar_state' not in st.session_state:
   st.session_state.sidebar_state = 'expanded'
   st.header('SOFTWARE_DEVELOPER_AI', divider='red')
   st.write("Concrete is the most important material in civil engineering.The concrete compressive strength is a highly nonlinear function of age andingredients. These ingredients include cement, blast furnace slag, fly ash, water, superplasticizer, coarse aggregate, and fine aggregate.So this the prediction for the strength regarding to the dependent features")
   st.write("-- Input Variable --")
   st.write("Age in Year")
   st.write("Na_to_K")
   st.write("Sex         --> Male OR Female")
   st.write("BP          --> HIGH - LOW - NORMAL")
   st.write("Cholesterol --> HIGH - NORMAL")
   st.write("-- Output Variable --")
   st.write("DRUG_TYPE   -->[ DRUG_Y  , DRUG_B , DRUG_C , DRUG_A , DRUG_X]")
   st.header('_MAY_BE_LIFE_ is :blue[cool] :sunglasses:')
   st.write('-------------------------------CONCLUSOR---------------------------------')
  data_matrix = [['ITEM', 'DESCIBTION'],
               ['DATA_SET', 'CONCRETE_STRENGTH'],
               ['MODEL', 'XGBOOST_REGRESSION'],
               ['SCORE_TRAIN-%', 99.58],
               ['SCORE_TEST-%', 91.04]]
  fig = ff.create_table(data_matrix)
  st.plotly_chart(fig)



"""<a style='text-decoration:none;line-height:16px;display:flex;color:#5B5B62;padding:10px;justify-content:end;' href='https://deepnote.com?utm_source=created-in-deepnote-cell&projectId=ba4b953b-0886-4d35-a77c-85c6bab3f5e8' target="_blank">
<img alt='Created in deepnote.com' style='display:inline;max-height:16px;margin:0px;margin-right:7.5px;' src='data:image/svg+xml;base64,PD94bWwgdmVyc2lvbj0iMS4wIiBlbmNvZGluZz0iVVRGLTgiPz4KPHN2ZyB3aWR0aD0iODBweCIgaGVpZ2h0PSI4MHB4IiB2aWV3Qm94PSIwIDAgODAgODAiIHZlcnNpb249IjEuMSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIiB4bWxuczp4bGluaz0iaHR0cDovL3d3dy53My5vcmcvMTk5OS94bGluayI+CiAgICA8IS0tIEdlbmVyYXRvcjogU2tldGNoIDU0LjEgKDc2NDkwKSAtIGh0dHBzOi8vc2tldGNoYXBwLmNvbSAtLT4KICAgIDx0aXRsZT5Hcm91cCAzPC90aXRsZT4KICAgIDxkZXNjPkNyZWF0ZWQgd2l0aCBTa2V0Y2guPC9kZXNjPgogICAgPGcgaWQ9IkxhbmRpbmciIHN0cm9rZT0ibm9uZSIgc3Ryb2tlLXdpZHRoPSIxIiBmaWxsPSJub25lIiBmaWxsLXJ1bGU9ImV2ZW5vZGQiPgogICAgICAgIDxnIGlkPSJBcnRib2FyZCIgdHJhbnNmb3JtPSJ0cmFuc2xhdGUoLTEyMzUuMDAwMDAwLCAtNzkuMDAwMDAwKSI+CiAgICAgICAgICAgIDxnIGlkPSJHcm91cC0zIiB0cmFuc2Zvcm09InRyYW5zbGF0ZSgxMjM1LjAwMDAwMCwgNzkuMDAwMDAwKSI+CiAgICAgICAgICAgICAgICA8cG9seWdvbiBpZD0iUGF0aC0yMCIgZmlsbD0iIzAyNjVCNCIgcG9pbnRzPSIyLjM3NjIzNzYyIDgwIDM4LjA0NzY2NjcgODAgNTcuODIxNzgyMiA3My44MDU3NTkyIDU3LjgyMTc4MjIgMzIuNzU5MjczOSAzOS4xNDAyMjc4IDMxLjY4MzE2ODMiPjwvcG9seWdvbj4KICAgICAgICAgICAgICAgIDxwYXRoIGQ9Ik0zNS4wMDc3MTgsODAgQzQyLjkwNjIwMDcsNzYuNDU0OTM1OCA0Ny41NjQ5MTY3LDcxLjU0MjI2NzEgNDguOTgzODY2LDY1LjI2MTk5MzkgQzUxLjExMjI4OTksNTUuODQxNTg0MiA0MS42NzcxNzk1LDQ5LjIxMjIyODQgMjUuNjIzOTg0Niw0OS4yMTIyMjg0IEMyNS40ODQ5Mjg5LDQ5LjEyNjg0NDggMjkuODI2MTI5Niw0My4yODM4MjQ4IDM4LjY0NzU4NjksMzEuNjgzMTY4MyBMNzIuODcxMjg3MSwzMi41NTQ0MjUgTDY1LjI4MDk3Myw2Ny42NzYzNDIxIEw1MS4xMTIyODk5LDc3LjM3NjE0NCBMMzUuMDA3NzE4LDgwIFoiIGlkPSJQYXRoLTIyIiBmaWxsPSIjMDAyODY4Ij48L3BhdGg+CiAgICAgICAgICAgICAgICA8cGF0aCBkPSJNMCwzNy43MzA0NDA1IEwyNy4xMTQ1MzcsMC4yNTcxMTE0MzYgQzYyLjM3MTUxMjMsLTEuOTkwNzE3MDEgODAsMTAuNTAwMzkyNyA4MCwzNy43MzA0NDA1IEM4MCw2NC45NjA0ODgyIDY0Ljc3NjUwMzgsNzkuMDUwMzQxNCAzNC4zMjk1MTEzLDgwIEM0Ny4wNTUzNDg5LDc3LjU2NzA4MDggNTMuNDE4MjY3Nyw3MC4zMTM2MTAzIDUzLjQxODI2NzcsNTguMjM5NTg4NSBDNTMuNDE4MjY3Nyw0MC4xMjg1NTU3IDM2LjMwMzk1NDQsMzcuNzMwNDQwNSAyNS4yMjc0MTcsMzcuNzMwNDQwNSBDMTcuODQzMDU4NiwzNy43MzA0NDA1IDkuNDMzOTE5NjYsMzcuNzMwNDQwNSAwLDM3LjczMDQ0MDUgWiIgaWQ9IlBhdGgtMTkiIGZpbGw9IiMzNzkzRUYiPjwvcGF0aD4KICAgICAgICAgICAgPC9nPgogICAgICAgIDwvZz4KICAgIDwvZz4KPC9zdmc+' > </img>
Created in <span style='font-weight:600;margin-left:4px;'>Deepnote</span></a>
"""
